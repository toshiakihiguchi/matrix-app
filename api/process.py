import os
import io
import json
import re
import base64
import urllib.parse
from http.server import BaseHTTPRequestHandler
from google import genai
from google.genai import types
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ReportLab (PDF生成用) のインポート
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

# 日本語フォント登録（HeiseiKakuGo-W5）
pdfmetrics.registerFont(UnicodeCIDFont('HeiseiKakuGo-W5'))

class handler(BaseHTTPRequestHandler):
    def do_POST(self):
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body.decode('utf-8'))

            action = data.get("action", "all")
            dept_time = data.get("deptTime", "明日の 10:30 出発")

            # -------------------------------------------------------------
            # STEP 1: AI解析リクエスト (action == 'parse' または 旧構成 'all')
            # -------------------------------------------------------------
            if action in ["parse", "all"]:
                file_b64 = data.get("file")
                mime_type = data.get("mimeType", "image/png")

                if not file_b64:
                    self.send_error(400, "ファイルデータが送信されていません。")
                    return

                api_key = os.environ.get("GEMINI_API_KEY")
                if not api_key:
                    self.send_error(500, "Vercelの環境変数 GEMINI_API_KEY が設定されていません。")
                    return

                client = genai.Client(api_key=api_key)

                prompt = (
                    "添付資料（画像またはPDF）に記載されている全ての施設名・店舗名・住所・駅名を読み取り、"
                    "以下の【略称変換ルール】を厳格に適用して正式名称を特定し、"
                    f"「{dept_time}」における移動所要時間マトリックスデータをJSONのみで作成してください。\n\n"
                    "【略称変換ルール】\n"
                    "- みらい千林西 ➔ 関西みらい銀行 千林西支店\n"
                    "- 関西みらい 出来島 ➔ 関西みらい銀行 出来島支店\n"
                    "- フロンティア 南大阪 ➔ フロンティア不動産販売 南大阪店\n"
                    "- 池田泉州銀行 交野支店 ➔ 池田泉州銀行 交野支店（最寄り：JR星田駅）\n"
                    "- ろうきん / 近畿労金 ➔ 近畿労働金庫\n"
                    "- だいしん / 大阪信金 ➔ 大阪信用金庫\n"
                    "- 京信 ➔ 京都信用金庫\n"
                    "- 中信 ➔ 京都中央信用金庫\n"
                    "- あましん ➔ 尼崎信用金庫\n"
                    "- ひまわり ➔ 兵庫ひまわり信用組合\n"
                    "- ひょうしん ➔ 兵庫信用金庫\n"
                    "- 京滋 ➔ 京滋信用組合\n"
                    "- 京銀 ➔ 京都銀行\n"
                    "- ひらしん ➔ 枚方信用金庫\n"
                    "- ドリーム ➔ ドリームホーム\n"
                    "- ドリーム本社 ➔ dreamtown本社\n"
                    "- ドリーム洛西口 ➔ ドリームホーム 洛西口駅前店\n\n"
                    "【出力JSON構造】\n"
                    "{\n"
                    '  "locations": [\n'
                    '    { "no": 1, "raw_name": "画像表記", "official_name": "正式名称", "address": "住所", "stations": "最寄り駅一覧" }\n'
                    "  ],\n"
                    '  "transit": [\n'
                    '    ["出発地／目的地", "拠点A", "拠点B"],\n'
                    '    ["拠点A", "同地点", "所要時間・乗換"],\n'
                    '    ["拠点B", "所要時間・乗換", "同地点"]\n'
                    "  ],\n"
                    '  "driving": [\n'
                    '    ["出発地／目的地", "拠点A", "拠点B"],\n'
                    '    ["拠点A", "同地点", "〇分 (〇km)"],\n'
                    '    ["拠点B", "〇分 (〇km)", "同地点"]\n'
                    "  ],\n"
                    '  "walking": [\n'
                    '    ["出発地／目的地", "拠点A", "拠点B"],\n'
                    '    ["拠点A", "同地点", "〇分 (〇km)"],\n'
                    '    ["拠点B", "〇分 (〇km)", "同地点"]\n'
                    "  ]\n"
                    "}\n"
                )

                file_bytes = base64.b64decode(file_b64)
                file_part = types.Part.from_bytes(data=file_bytes, mime_type=mime_type)
                
                response = client.models.generate_content(
                    model='gemini-3.6-flash',
                    contents=[prompt, file_part]
                )
                
                res_text = response.text.strip()
                json_match = re.search(r'\{.*\}', res_text, re.DOTALL)
                clean_json_str = json_match.group(0) if json_match else res_text
                result_data = json.loads(clean_json_str)

                # action == 'parse' の場合は画面側へJSONデータを返却
                if action == "parse":
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json; charset=utf-8')
                    self.end_headers()
                    self.wfile.write(json.dumps(result_data, ensure_ascii=False).encode('utf-8'))
                    return

            # -------------------------------------------------------------
            # STEP 2: ファイル出力リクエスト (action == 'export')
            # -------------------------------------------------------------
            if action == "export":
                result_data = data.get("resultData", {})
                output_format = data.get("format", "excel")
            else:
                output_format = "excel"

            # ==========================================
            # A. PDF出力処理 (ReportLab)
            # ==========================================
            if output_format == "pdf":
                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(
                    pdf_buffer,
                    pagesize=landscape(A4),
                    rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
                )

                elements = []
                styles = getSampleStyleSheet()
                
                title_style = ParagraphStyle(
                    'JPTitle', fontName='HeiseiKakuGo-W5', fontSize=12, leading=16, textColor=colors.HexColor('#1F4E78')
                )
                cell_style = ParagraphStyle(
                    'JPCell', fontName='HeiseiKakuGo-W5', fontSize=7.5, leading=9, alignment=1
                )
                header_style = ParagraphStyle(
                    'JPHeader', fontName='HeiseiKakuGo-W5', fontSize=8, leading=10, alignment=1, textColor=colors.white
                )
                map_style = ParagraphStyle(
                    'JPMapCell', fontName='HeiseiKakuGo-W5', fontSize=7.5, leading=9, alignment=0
                )

                modes = [("公共交通機関", "transit", "transit"), ("車利用", "driving", "driving"), ("徒歩利用", "walking", "walking")]

                for idx, (mode_title, data_key, mode_param) in enumerate(modes):
                    if idx > 0:
                        elements.append(PageBreak())

                    elements.append(Paragraph(f"<b>■ ROUTE MATRIX ({mode_title}) - 出発想定：{dept_time}</b>", title_style))
                    elements.append(Spacer(1, 10))

                    matrix = result_data.get(data_key, [])
                    if matrix:
                        table_data = []
                        # ヘッダー行
                        header_row = [Paragraph("出発地 ＼ 目的地", header_style)]
                        for col_name in matrix[0][1:]:
                            header_row.append(Paragraph(col_name, header_style))
                        table_data.append(header_row)

                        # データ行
                        for r_data in matrix[1:]:
                            origin_name = r_data[0]
                            row_cells = [Paragraph(f"<b>{origin_name}</b>", ParagraphStyle('JPOrigin', fontName='HeiseiKakuGo-W5', fontSize=8, leading=10, alignment=1))]
                            
                            for c_idx, val in enumerate(r_data[1:], start=1):
                                dest_name = matrix[0][c_idx]
                                if origin_name == dest_name or val in ["同地点", "-"]:
                                    row_cells.append(Paragraph("-", cell_style))
                                else:
                                    maps_url = f"https://www.google.com/maps/dir/?api=1&origin={urllib.parse.quote(origin_name)}&destination={urllib.parse.quote(dest_name)}&travelmode={mode_param}"
                                    link_html = f'<a href="{maps_url}"><font color="#0000FF"><u>{val}</u></font></a>'
                                    row_cells.append(Paragraph(link_html, cell_style))
                            table_data.append(row_cells)

                        num_cols = len(matrix[0])
                        col_width = 800 / num_cols
                        
                        t = Table(table_data, colWidths=[col_width]*num_cols)
                        t.setStyle(TableStyle([
                            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
                            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F2F5F8')),
                            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
                            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                        ]))
                        elements.append(t)

                # 全体マップページ追加
                elements.append(PageBreak())
                elements.append(Paragraph("<b>■ 全体マップ・施設一覧</b>", title_style))
                elements.append(Spacer(1, 10))

                map_headers = ["No.", "画像上の表記", "特定された正式店舗・施設名", "正式住所", "Googleマップ", "利用可能な最寄り駅一覧"]
                map_table_data = [[Paragraph(h, header_style) for h in map_headers]]

                for loc in result_data.get("locations", []):
                    official = loc.get("official_name", "")
                    map_url = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(official)}"
                    
                    row = [
                        Paragraph(str(loc.get("no", "")), cell_style),
                        Paragraph(str(loc.get("raw_name", "")), map_style),
                        Paragraph(official, map_style),
                        Paragraph(str(loc.get("address", "")), map_style),
                        Paragraph(f'<a href="{map_url}"><font color="#0000FF"><u>マップ表示</u></font></a>', cell_style),
                        Paragraph(str(loc.get("stations", "")), map_style)
                    ]
                    map_table_data.append(row)

                t_map = Table(map_table_data, colWidths=[30, 110, 160, 220, 80, 200])
                t_map.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E78')),
                    ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#D9D9D9')),
                    ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
                ]))
                elements.append(t_map)

                doc.build(elements)
                pdf_data = pdf_buffer.getvalue()

                self.send_response(200)
                self.send_header('Content-Type', 'application/pdf')
                self.send_header('Content-Disposition', 'attachment; filename="matrix.pdf"')
                self.end_headers()
                self.wfile.write(pdf_data)
                return

            # ==========================================
            # B. Excel出力処理 (openpyxl)
            # ==========================================
            wb = openpyxl.Workbook()
            header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            header_font = Font(name="Meiryo", size=10, bold=True, color="FFFFFF")
            body_font = Font(name="Meiryo", size=9.5)
            link_font = Font(name="Meiryo", size=9.5, color="004B91", underline="single")
            
            thin_border = Border(
                left=Side(style='thin', color='D9D9D9'),
                right=Side(style='thin', color='D9D9D9'),
                top=Side(style='thin', color='D9D9D9'),
                bottom=Side(style='thin', color='D9D9D9')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            modes = [("公共交通機関", "transit", "transit"), ("車利用", "driving", "driving"), ("徒歩利用", "walking", "walking")]
            
            for idx, (title, data_key, mode_param) in enumerate(modes):
                ws = wb.active if idx == 0 else wb.create_sheet(title=title)
                if idx == 0:
                    ws.title = title
                
                ws.freeze_panes = 'B5'
                ws['A1'] = f"■ 出発想定日時：{dept_time}"
                ws['A1'].font = Font(name="Meiryo", bold=True, size=11, color="1F4E78")
                ws.row_dimensions[1].height = 24
                
                ws['A4'] = "出発地 ＼ 目的地"
                ws['A4'].fill = header_fill
                ws['A4'].font = header_font
                ws['A4'].alignment = center_align
                ws.row_dimensions[4].height = 40

                matrix_data = result_data.get(data_key, [])
                if matrix_data:
                    max_a_len = 16
                    for c_idx, col_name in enumerate(matrix_data[0][1:], start=2):
                        cell = ws.cell(row=4, column=c_idx, value=col_name)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = center_align
                        ws.column_dimensions[get_column_letter(c_idx)].width = 16

                    for r_idx, row_data in enumerate(matrix_data[1:], start=5):
                        ws.row_dimensions[r_idx].height = 32
                        origin_name = row_data[0]
                        max_a_len = max(max_a_len, len(str(origin_name)))
                        
                        a_cell = ws.cell(row=r_idx, column=1, value=origin_name)
                        a_cell.font = Font(name="Meiryo", size=9.5, bold=True)
                        a_cell.fill = PatternFill(start_color="F2F5F8", fill_type="solid")
                        a_cell.alignment = center_align
                        a_cell.border = thin_border

                        for c_idx, val in enumerate(row_data[1:], start=2):
                            dest_name = matrix_data[0][c_idx-1]
                            cell = ws.cell(row=r_idx, column=c_idx)
                            cell.alignment = center_align
                            cell.border = thin_border

                            if origin_name == dest_name or val in ["同地点", "-"]:
                                cell.value = "-"
                                cell.font = body_font
                            else:
                                cell.value = str(val)
                                maps_url = f"https://www.google.com/maps/dir/?api=1&origin={urllib.parse.quote(origin_name)}&destination={urllib.parse.quote(dest_name)}&travelmode={mode_param}"
                                cell.hyperlink = maps_url
                                cell.font = link_font

                    ws.column_dimensions['A'].width = max(max_a_len * 2.2, 22)

            # 全体マップシート
            ws_map = wb.create_sheet(title="全体マップ")
            ws_map.freeze_panes = 'A2'
            headers_map = ["No.", "画像上の表記", "特定された正式店舗・施設名", "正式住所", "Googleマップ", "利用可能な最寄り駅一覧と所要時間"]
            ws_map.append(headers_map)
            ws_map.row_dimensions[1].height = 28
            
            for c_idx, h_text in enumerate(headers_map, start=1):
                cell = ws_map.cell(row=1, column=c_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            col_widths = [6, 18, 26, 38, 16, 35]

            for loc in result_data.get("locations", []):
                official = loc.get("official_name", "")
                map_url = f"https://www.google.com/maps/search/?api=1&query={urllib.parse.quote(official)}"
                
                row_cells = [
                    loc.get("no"),
                    loc.get("raw_name"),
                    official,
                    loc.get("address"),
                    "マップを開く",
                    loc.get("stations")
                ]
                ws_map.append(row_cells)
                last_row = ws_map.max_row
                ws_map.row_dimensions[last_row].height = 26
                
                for c_idx in range(1, 7):
                    cell = ws_map.cell(row=last_row, column=c_idx)
                    cell.border = thin_border
                    cell.font = body_font
                    if c_idx in [1, 5]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                map_cell = ws_map.cell(row=last_row, column=5)
                map_cell.hyperlink = map_url
                map_cell.font = link_font

            for idx, width in enumerate(col_widths, start=1):
                ws_map.column_dimensions[get_column_letter(idx)].width = width

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_data = excel_buffer.getvalue()

            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="matrix.xlsx"')
            self.end_headers()
            self.wfile.write(excel_data)

        except Exception as e:
            self.send_response(500)
            self.send_header('Content-Type', 'text/plain; charset=utf-8')
            self.end_headers()
            self.wfile.write(f"処理エラーが発生しました:\n{str(e)}".encode('utf-8'))
